from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Optional, Union

from openpyxl import load_workbook

from etf_config import ETF_BY_CODE
from etf_matching import canonical_name, infer_etf_code


SUMMARY_HEADERS = {"乖离率", "状态诊断", "操作建议"}


def date_label(day: date, suffix: str = "收盘") -> str:
    return f"{day.month}.{day.day}{suffix}"


def _to_float(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, str):
        value = value.replace(",", "").replace("%", "").strip()
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _price_row_by_code(ws):
    mapping = {}
    for row in range(2, ws.max_row + 1):
        code = str(ws.cell(row, 2).value or "").strip()
        if code:
            mapping[code] = row
    return mapping


def _find_or_create_price_column(ws, label: str) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == label:
            return col

    insert_col = ws.max_column + 1
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value in SUMMARY_HEADERS:
            insert_col = col
            break

    ws.insert_cols(insert_col)
    ws.cell(1, insert_col).value = label
    return insert_col


def update_workbook(
    excel_path: Union[str, Path],
    output_path: Union[str, Path],
    day: date,
    prices: list[dict],
    holdings: list[dict],
    cash_balance: Optional[float],
) -> Path:
    wb = load_workbook(excel_path)
    update_price_sheet(wb, day, prices)
    append_position_block(wb, day, holdings, cash_balance)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def update_price_sheet(wb, day: date, prices: list[dict]) -> None:
    ws = wb["股价跟进表"]
    col = _find_or_create_price_column(ws, date_label(day, "收盘"))
    rows = _price_row_by_code(ws)

    for item in prices:
        code = infer_etf_code(item.get("name"), item.get("code"))
        close = _to_float(item.get("close"))
        if not code or close is None:
            continue
        row = rows.get(code)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row, 1).value = row - 1
            ws.cell(row, 2).value = code
            ws.cell(row, 3).value = canonical_name(code, item.get("name"))
            rows[code] = row
        ws.cell(row, col).value = close


def append_position_block(wb, day: date, holdings: list[dict], cash_balance: Optional[float]) -> None:
    ws = wb["仓位图"]
    label = f"日期{day.month}.{day.day}"
    start_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == label:
            start_row = row
            break

    if start_row is None:
        start_row = ws.max_row + 2
    else:
        next_block = ws.max_row + 1
        for row in range(start_row + 1, ws.max_row + 1):
            value = ws.cell(row, 1).value
            if isinstance(value, str) and value.startswith("日期"):
                next_block = row
                break
        ws.delete_rows(start_row, next_block - start_row)

    rows = [[label, None, None, None, None, None, None]]
    rows.append(["名称", "ETF代码", "持有市值", "持仓盈亏", "持仓盈亏比例", "个股仓位", "持仓数量"])

    for item in holdings:
        code = infer_etf_code(item.get("name"), item.get("code"))
        if not code:
            continue
        rows.append(
            [
                canonical_name(code, item.get("name")),
                code,
                _to_float(item.get("market_value")),
                _to_float(item.get("pnl")),
                _normalize_pct(item.get("pnl_pct")),
                _normalize_pct(item.get("weight")),
                _to_float(item.get("quantity")),
            ]
        )

    if cash_balance is not None:
        rows.append(["现金", "-", _to_float(cash_balance), "-", None, "-", "-"])

    for r_offset, values in enumerate(rows):
        for c_offset, value in enumerate(values):
            ws.cell(start_row + r_offset, 1 + c_offset).value = value


def _normalize_pct(value):
    number = _to_float(value)
    if number is None:
        return None
    if abs(number) > 1:
        return number / 100
    return number
