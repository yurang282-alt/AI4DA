from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import time
from datetime import date
from pathlib import Path
from typing import Any

import sys
from dotenv import load_dotenv
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from etf_matching import canonical_name, infer_etf_code


PRICE_FIELDS = ["ETF代码", "日期", "ETF名称", "收盘价", "涨跌幅%"]
HOLDING_FIELDS = [
    "ETF代码",
    "日期",
    "ETF名称",
    "持有市值",
    "持仓盈亏",
    "盈亏比例%",
    "个股仓位%",
    "持仓数量",
    "可用数量",
    "现金余额",
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--year", type=int, default=2026)
    args = parser.parse_args()

    load_dotenv(".env")
    app_token = os.environ["FEISHU_BITABLE_APP_TOKEN"]
    price_table_id = os.environ["FEISHU_PRICE_TABLE_ID"]
    holding_table_id = os.environ["FEISHU_HOLDING_TABLE_ID"]

    prices, holdings = parse_excel(Path(args.excel), args.year)
    print(f"Parsed price records: {len(prices)}")
    print(f"Parsed holding records: {len(holdings)}")

    if args.dry_run:
        print(json.dumps({"prices_sample": prices[:3], "holdings_sample": holdings[:3]}, ensure_ascii=False, indent=2))
        return

    batch_create(app_token, price_table_id, PRICE_FIELDS, prices)
    batch_create(app_token, holding_table_id, HOLDING_FIELDS, holdings)
    print("Import completed.")


def parse_excel(path: Path, year: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    return parse_prices(wb["股价跟进表"], year), parse_holdings(wb["仓位图"], year)


def parse_prices(ws, year: int) -> list[dict[str, Any]]:
    close_columns = []
    for col in range(4, ws.max_column + 1):
        header = ws.cell(1, col).value
        if isinstance(header, str) and re.match(r"^\d+\.\d+收盘$", header):
            close_columns.append((col, parse_cn_month_day(header.replace("收盘", ""), year)))

    records = []
    for row in range(2, ws.max_row + 1):
        code = infer_etf_code(ws.cell(row, 3).value, ws.cell(row, 2).value)
        if not code:
            continue
        for col, day in close_columns:
            close = to_float(ws.cell(row, col).value)
            if close is None:
                continue
            records.append(
                {
                    "ETF代码": code,
                    "日期": day.isoformat(),
                    "ETF名称": canonical_name(code, ws.cell(row, 3).value),
                    "收盘价": close,
                    "涨跌幅%": None,
                }
            )
    return records


def parse_holdings(ws, year: int) -> list[dict[str, Any]]:
    block_starts = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and value.startswith("日期"):
            block_starts.append((row, parse_cn_month_day(value.replace("日期", ""), year)))

    records = []
    for index, (start, day) in enumerate(block_starts):
        end = block_starts[index + 1][0] if index + 1 < len(block_starts) else ws.max_row + 1
        for row in range(start + 2, end):
            name = ws.cell(row, 1).value
            if name in (None, ""):
                continue
            if str(name).strip() == "现金":
                records.append(
                    {
                        "ETF代码": "CASH",
                        "日期": day.isoformat(),
                        "ETF名称": "现金",
                        "持有市值": None,
                        "持仓盈亏": None,
                        "盈亏比例%": None,
                        "个股仓位%": None,
                        "持仓数量": None,
                        "可用数量": None,
                        "现金余额": to_float(ws.cell(row, 3).value),
                    }
                )
                continue

            code = infer_etf_code(name, ws.cell(row, 2).value)
            if not code:
                print(f"Skip unmatched holding row {row}: {name}")
                continue
            records.append(
                {
                    "ETF代码": code,
                    "日期": day.isoformat(),
                    "ETF名称": canonical_name(code, name),
                    "持有市值": to_float(ws.cell(row, 3).value),
                    "持仓盈亏": to_float(ws.cell(row, 4).value),
                    "盈亏比例%": pct_to_number(ws.cell(row, 5).value),
                    "个股仓位%": pct_to_number(ws.cell(row, 6).value),
                    "持仓数量": to_float(ws.cell(row, 7).value),
                    "可用数量": None,
                    "现金余额": None,
                }
            )
    return records


def parse_cn_month_day(value: str, year: int) -> date:
    month, day = value.split(".", 1)
    return date(year, int(month), int(day))


def to_float(value: Any) -> float | None:
    if value in (None, "", "-"):
        return None
    if isinstance(value, str):
        value = value.replace(",", "").replace("%", "").strip()
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def pct_to_number(value: Any) -> float | None:
    number = to_float(value)
    if number is None:
        return None
    if abs(number) <= 1:
        return number * 100
    return number


def batch_create(app_token: str, table_id: str, fields: list[str], records: list[dict[str, Any]]) -> None:
    for start in range(0, len(records), 100):
        chunk = records[start : start + 100]
        payload = {"fields": fields, "rows": [[record.get(field) for field in fields] for record in chunk]}
        command = [
            "lark-cli",
            "base",
            "+record-batch-create",
            "--base-token",
            app_token,
            "--table-id",
            table_id,
            "--json",
            json.dumps(payload, ensure_ascii=False),
            "--as",
            "user",
        ]
        result = subprocess.run(command, capture_output=True, text=True, check=False)
        if result.returncode != 0:
            raise RuntimeError(result.stderr or result.stdout)
        response = json.loads(result.stdout)
        if not response.get("ok"):
            raise RuntimeError(result.stdout)
        print(f"Imported {min(start + len(chunk), len(records))}/{len(records)} into {table_id}")
        time.sleep(0.8)


if __name__ == "__main__":
    main()
